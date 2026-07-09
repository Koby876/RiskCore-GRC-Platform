"""
ui/reports.py
──────────────
Export & Report Page — Phase 7 migration.

Three large action cards (Image 6 layout):
  📄 Full GRC PDF Report
  📊 Export CSV
  ◧  Backup Database

Plus:
  Register count + Company Name + Classification header bar
  Database Backup status section
  Report Includes checklist (10 items)

PDF generation runs on a QThread so the UI never freezes
during the ReportLab render.
"""

from __future__ import annotations

import csv
import datetime
from pathlib import Path

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QScrollArea, QLineEdit, QComboBox, QGridLayout,
    QFileDialog, QMessageBox, QSizePolicy,
)
from PySide6.QtCore import Qt, Signal, QThread, QObject
from PySide6.QtGui import QFont, QCursor

from assets.themes.design_system import Colors, Fonts, Spacing, Radius
from core.database.db import (
    get_risks, get_organisation_scope,
    backup_database, audit, today,
    BASE_DIR, load_settings,
)
from core.services.ai_service import (
    build_data_driven_analysis, generate_pdf_report,
)


# ── PDF generation worker ─────────────────────────────────────────────────────

class PdfWorker(QObject):
    finished = Signal(str)   # output path on success
    error    = Signal(str)

    def __init__(self, analysis, risks, company,
                 path, clf, org_scope):
        super().__init__()
        self._analysis  = analysis
        self._risks     = risks
        self._company   = company
        self._path      = path
        self._clf       = clf
        self._org_scope = org_scope

    def run(self) -> None:
        try:
            generate_pdf_report(
                self._analysis, self._risks,
                self._company, self._path,
                self._clf, org_scope=self._org_scope)
            audit("EXPORT_PDF",
                  detail=(f"{len(self._risks)} risks → "
                          f"{Path(self._path).name}"))
            self.finished.emit(self._path)
        except Exception as e:
            self.error.emit(str(e))


# ── Helpers ───────────────────────────────────────────────────────────────────

def _lbl(text, font=None,
         color=Colors.TEXT_MUTED) -> QLabel:
    l = QLabel(str(text))
    l.setFont(font or Fonts.label())
    l.setStyleSheet(f"color: {color}; border: none;")
    return l


def _entry(placeholder="", width=240) -> QLineEdit:
    e = QLineEdit()
    e.setPlaceholderText(placeholder)
    e.setFixedHeight(32)
    e.setFixedWidth(width)
    e.setStyleSheet(f"""
        QLineEdit {{
            background-color: {Colors.BG_CARD2};
            color: {Colors.TEXT_PRIMARY};
            border: 1px solid {Colors.BG_BORDER};
            border-radius: {Radius.SM}px;
            padding: 0 10px; font-size: 10pt;
        }}
    """)
    return e


def _combo(items, width=160) -> QComboBox:
    c = QComboBox()
    c.addItems(items)
    c.setFixedWidth(width)
    c.setFixedHeight(32)
    c.setStyleSheet(f"""
        QComboBox {{
            background-color: {Colors.BG_CARD2};
            color: {Colors.TEXT_PRIMARY};
            border: 1px solid {Colors.BG_BORDER};
            border-radius: {Radius.SM}px;
            padding: 2px 8px; font-size: 10pt;
        }}
        QComboBox QAbstractItemView {{
            background-color: {Colors.BG_CARD};
            color: {Colors.TEXT_PRIMARY};
            selection-background-color: {Colors.ACCENT_BLUE};
        }}
    """)
    return c


# ── Reports Page ──────────────────────────────────────────────────────────────

class ReportsPage(QWidget):
    """
    Export & Report page.

    Accepts an optional cached analysis dict so that if the user
    has run AI analysis the same session, the PDF uses that result
    rather than building a data-driven fallback.

    Signals
    -------
    navigate(str) : page navigation request
    """
    navigate = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._cached_analysis: dict | None = None
        self._settings = load_settings()
        self._setup_ui()

    def set_analysis(self, analysis: dict) -> None:
        """Called by main_window when AI analysis completes."""
        self._cached_analysis = analysis

    def _org_name_from_settings(self) -> str:
        """Single source of truth — reads org name from scope then settings."""
        try:
            scope = get_organisation_scope()
            if scope and scope.get("organisation_name"):
                return scope["organisation_name"]
            s = self._settings or load_settings() or {}
            return s.get("organisation_name", "Your Organisation")
        except Exception:
            return "Your Organisation"

    def _reset_company_name(self) -> None:
        """Sync company field back to saved org name from Settings."""
        if hasattr(self, "_company_input"):
            self._company_input.setText(self._org_name_from_settings())

    def refresh(self) -> None:
        """Refresh risk cache, org name, classification, and backup status."""
        from core.database.db import get_risks as _gr
        self._risks_cache = [dict(r) for r in _gr()]
        cnt = len(self._risks_cache)
        if hasattr(self, '_cnt_lbl'):
            self._cnt_lbl.setText(f"Register contains {cnt} risk(s)")
        self._update_export_buttons()
        # Always sync company name from settings on refresh
        if hasattr(self, "_company_input"):
            self._company_input.setText(self._org_name_from_settings())
        # Sync classification from settings
        if hasattr(self, "_clf_combo"):
            self._settings = load_settings() or self._settings
            clf = self._settings.get("default_classification", "CONFIDENTIAL")
            self._clf_combo.setCurrentText(clf)
        # Refresh backup status labels
        self._refresh_backup_status()

    def _refresh_backup_status(self) -> None:
        """Update backup date and status labels from disk."""
        if not hasattr(self, "_bk_last_lbl"):
            return
        import datetime as _dt
        try:
            from riskcore_phase2 import BASE_DIR
            bk_dir = BASE_DIR / "backups"
            files = sorted(
                bk_dir.glob("riskcore_backup_*.db"),
                key=lambda p: p.stat().st_mtime, reverse=True)
            last_bk  = (_dt.datetime.fromtimestamp(
                files[0].stat().st_mtime).strftime("%Y-%m-%d  %H:%M")
                        if files else "Never yet")
            bk_status = "Healthy" if files else "No backup yet"
            bk_color  = Colors.SUCCESS_LT if files else Colors.CRITICAL
        except Exception:
            last_bk   = "Unknown"
            bk_status = "Unknown"
            bk_color  = Colors.TEXT_MUTED

        self._bk_last_lbl.setText(last_bk)
        self._bk_last_lbl.setStyleSheet(
            f"color: {Colors.ACCENT_BLUE if files else Colors.TEXT_MUTED}; border: none;")
        self._bk_status_lbl.setText(bk_status)
        self._bk_status_lbl.setStyleSheet(
            f"color: {bk_color}; border: none;")

    def _update_export_buttons(self) -> None:
        """Enable/disable export buttons based on risk count."""
        has_risks = len(self._risks_cache) > 0
        for btn in getattr(self, '_export_btns', []):
            btn.setEnabled(has_risks)
            btn.setStyleSheet(btn.property('_style_enabled')
                              if has_risks
                              else btn.property('_style_disabled'))

    # ── UI Construction ───────────────────────────────────────────────────────

    def _setup_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_header())

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        body = QWidget()
        body.setStyleSheet(
            f"background-color: {Colors.BG_DEEP};")
        self._body = QVBoxLayout(body)
        self._body.setContentsMargins(
            Spacing.XL, Spacing.MD, Spacing.XL, Spacing.XL)
        self._body.setSpacing(Spacing.MD)
        scroll.setWidget(body)
        root.addWidget(scroll, 1)

        self._body.addWidget(self._build_top_bar())

        # Status label (shown after actions)
        self._status_lbl = QLabel("")
        self._status_lbl.setFont(Fonts.label())
        self._status_lbl.setStyleSheet(
            f"color: {Colors.TEXT_MUTED};")
        self._body.addWidget(self._status_lbl)

        self._body.addWidget(self._build_action_cards())
        self._body.addWidget(self._build_backup_section())
        self._body.addWidget(self._build_includes())
        self._body.addStretch()

    def _build_header(self) -> QWidget:
        w = QWidget()
        w.setStyleSheet(f"background-color: {Colors.BG_DEEP};")
        hl = QHBoxLayout(w)
        hl.setContentsMargins(
            Spacing.XL, Spacing.LG, Spacing.XL, Spacing.SM)
        col = QVBoxLayout()
        col.setSpacing(2)
        t = QLabel("↗  Export & Report")
        t.setFont(Fonts.heading_1())
        t.setStyleSheet(f"color: {Colors.TEXT_PRIMARY};")
        col.addWidget(t)
        s = QLabel(
            "Industry-standard GRC report  ·  "
            "NIST SP 800-30 aligned")
        s.setFont(Fonts.label())
        s.setStyleSheet(f"color: {Colors.TEXT_MUTED};")
        col.addWidget(s)
        hl.addLayout(col)
        hl.addStretch()
        return w

    def _build_top_bar(self) -> QFrame:
        """Register count, company name, classification."""
        # Always query fresh — never rely on stale cache
        risks = get_risks()
        self._risks_cache = [dict(r) for r in risks]
        cnt = len(risks)

        bar = QFrame()
        bar.setStyleSheet(
            f"background-color: {Colors.BG_CARD};"
            f"border-radius: {Radius.LG}px;"
            f"border: 1px solid {Colors.BG_BORDER};")
        hl = QHBoxLayout(bar)
        hl.setContentsMargins(Spacing.LG, 0, Spacing.LG, 0)
        hl.setSpacing(Spacing.MD)

        self._cnt_lbl = QLabel(f"Register contains {cnt} risk(s)")
        self._cnt_lbl.setFont(QFont(Fonts.FAMILY, 13))
        self._cnt_lbl.setStyleSheet(
            f"color: {Colors.TEXT_PRIMARY}; border: none;")
        hl.addWidget(self._cnt_lbl, 1)

        hl.addWidget(_lbl("Company Name"))
        self._company_input = _entry("Your Organisation", 220)
        self._company_input.setText(self._org_name_from_settings())
        hl.addWidget(self._company_input)

        # Reset button — syncs company field back to saved org name
        reset_btn = QPushButton("↺")
        reset_btn.setFixedSize(28, 28)
        reset_btn.setFont(QFont(Fonts.FAMILY, 12))
        reset_btn.setToolTip(
            "Reset to Organisation Name from Settings\n"
            "Current: " + self._org_name_from_settings())
        reset_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        reset_btn.setStyleSheet(f"""
            QPushButton {{
                background: {Colors.BG_BORDER};
                color: {Colors.TEXT_MUTED};
                border: none;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background: {Colors.ACCENT_BLUE};
                color: white;
            }}
        """)
        reset_btn.clicked.connect(self._reset_company_name)
        hl.addWidget(reset_btn)

        hl.addWidget(_lbl("Classification"))
        self._clf_combo = _combo(
            ["CONFIDENTIAL","RESTRICTED","INTERNAL","PUBLIC"], 155)
        default_clf = self._settings.get(
            "default_classification", "CONFIDENTIAL")
        self._clf_combo.setCurrentText(default_clf)
        hl.addWidget(self._clf_combo)
        return bar

    def _build_action_cards(self) -> QWidget:
        """Three large coloured action cards — Image 6 layout."""
        row = QWidget()
        row.setStyleSheet("border: none;")
        hl = QHBoxLayout(row)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.setSpacing(Spacing.SM)

        def _card(icon, title, sub, color, cmd):
            f = QFrame()
            f.setStyleSheet(
                f"background-color: {color};"
                f"border-radius: {Radius.XL}px;"
                f"border: none;")
            vl = QVBoxLayout(f)
            vl.setContentsMargins(
                Spacing.LG, Spacing.LG,
                Spacing.LG, Spacing.LG)
            vl.setSpacing(Spacing.SM)

            icon_lbl = QLabel(icon)
            icon_lbl.setFont(QFont(Fonts.FAMILY, 30))
            icon_lbl.setStyleSheet("color: white; border: none;")
            vl.addWidget(icon_lbl)

            t_lbl = QLabel(title)
            t_lbl.setFont(
                QFont(Fonts.FAMILY, 14, QFont.Weight.Bold))
            t_lbl.setStyleSheet("color: white; border: none;")
            vl.addWidget(t_lbl)

            s_lbl = QLabel(sub)
            s_lbl.setFont(Fonts.label_sm())
            s_lbl.setStyleSheet(
                "color: rgba(255,255,255,0.8); border: none;")
            s_lbl.setWordWrap(True)
            vl.addWidget(s_lbl, 1)

            btn = QPushButton("Generate →")
            btn.setFixedHeight(36)
            btn.setCursor(
                QCursor(Qt.CursorShape.PointingHandCursor))
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: white;
                    color: {color};
                    border: none;
                    border-radius: {Radius.SM}px;
                    font-size: 12pt; font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: rgba(255,255,255,0.85);
                }}
            """)
            btn.clicked.connect(cmd)
            vl.addWidget(btn)
            hl.addWidget(f)

        _card("📄", "Full GRC PDF Report",
              "Comprehensive report with all risks, "
              "treatments and mappings",
              Colors.ACCENT_RED, self._export_pdf)
        _card("📊", "Export CSV",
              "Export risk register and treatments "
              "to CSV format",
              Colors.SUCCESS, self._export_csv)
        _card("📗", "Export Excel",
              "Styled workbook — Executive Summary, "
              "Risks, and Treatments sheets",
              Colors.FW_NIST, self._export_excel)
        _card("◧", "Backup Database",
              "Create a secure timestamped backup of "
              "your RiskCore database",
              Colors.ACCENT_BLUE, self._export_backup)
        return row

    def _build_backup_section(self) -> QFrame:
        card = QFrame()
        card.setStyleSheet(
            f"background-color: {Colors.BG_CARD};"
            f"border-radius: {Radius.LG}px;"
            f"border: 1px solid {Colors.BG_BORDER};")
        gl = QGridLayout(card)
        gl.setContentsMargins(
            Spacing.LG, Spacing.MD, Spacing.LG, Spacing.MD)
        gl.setSpacing(Spacing.SM)
        gl.setColumnStretch(1, 1)

        t = QLabel("◧  Database Backup")
        t.setFont(QFont(Fonts.FAMILY, 13, QFont.Weight.Bold))
        t.setStyleSheet(
            f"color: {Colors.TEXT_PRIMARY}; border: none;")
        gl.addWidget(t, 0, 0, 1, 4)

        gl.addWidget(_lbl("Backup folder"), 1, 0)
        bk_dir_lbl = QLabel(str(BASE_DIR / "backups"))
        bk_dir_lbl.setFont(Fonts.label_sm())
        bk_dir_lbl.setStyleSheet(
            f"background-color: {Colors.BG_CARD2};"
            f"color: {Colors.TEXT_MUTED};"
            f"border-radius: {Radius.SM}px;"
            f"padding: 4px 10px; border: none;")
        gl.addWidget(bk_dir_lbl, 1, 1, 1, 2)

        # Last backup
        bk_dir = BASE_DIR / "backups"
        try:
            files = sorted(
                bk_dir.glob("riskcore_backup_*.db"),
                key=lambda p: p.stat().st_mtime, reverse=True)
            last_bk = (datetime.datetime.fromtimestamp(
                files[0].stat().st_mtime).strftime(
                    "%Y-%m-%d  %H:%M")
                       if files else "Never yet")
            bk_status = "Healthy" if files else "No backup yet"
            bk_color  = Colors.SUCCESS_LT if files else Colors.CRITICAL
        except Exception:
            last_bk   = "Unknown"
            bk_status = "Unknown"
            bk_color  = Colors.TEXT_MUTED

        gl.addWidget(_lbl("Last backup"), 2, 0)
        self._bk_last_lbl = QLabel(last_bk)
        self._bk_last_lbl.setFont(Fonts.label_sm())
        self._bk_last_lbl.setStyleSheet(
            f"color: {Colors.ACCENT_BLUE}; border: none;")
        gl.addWidget(self._bk_last_lbl, 2, 1)

        gl.addWidget(_lbl("Backup status"), 3, 0)
        self._bk_status_lbl = QLabel(bk_status)
        self._bk_status_lbl.setFont(Fonts.label_sm_bold())
        self._bk_status_lbl.setStyleSheet(
            f"color: {bk_color}; border: none;")
        gl.addWidget(self._bk_status_lbl, 3, 1)

        note = QLabel(
            "ⓘ  Each backup is a complete timestamped copy of "
            "riskcore.db. No schema changes are made.")
        note.setFont(Fonts.label_sm())
        note.setStyleSheet(
            f"color: {Colors.TEXT_DIM}; border: none;")
        note.setWordWrap(True)
        gl.addWidget(note, 4, 0, 1, 4)
        return card

    def _build_includes(self) -> QFrame:
        """Report includes checklist — 4-column grid."""
        card = QFrame()
        card.setStyleSheet(
            f"background-color: {Colors.BG_CARD};"
            f"border-radius: {Radius.LG}px;"
            f"border: 1px solid {Colors.BG_BORDER};")
        vl = QVBoxLayout(card)
        vl.setContentsMargins(
            Spacing.LG, Spacing.MD, Spacing.LG, Spacing.MD)
        vl.setSpacing(Spacing.SM)

        t = QLabel("✓  Report includes")
        t.setFont(QFont(Fonts.FAMILY, 13, QFont.Weight.Bold))
        t.setStyleSheet(
            f"color: {Colors.TEXT_PRIMARY}; border: none;")
        vl.addWidget(t)

        items = [
            "Executive Summary",
            "Risk Register",
            "Risk Heat Map",
            "Treatment Plan",
            "Framework Mappings",
            "NIST CSF, ISO 27001, MITRE ATT&CK",
            "Risk Scoring (NIST SP 800-30)",
            "Recommendations",
            "Methodology",
            "Appendices",
        ]
        grid = QWidget()
        gl = QGridLayout(grid)
        gl.setContentsMargins(0, 0, 0, 0)
        gl.setSpacing(4)
        for i in range(4):
            gl.setColumnStretch(i, 1)
        for idx, item in enumerate(items):
            col = idx % 4
            row = idx // 4
            l = QLabel(f"✓  {item}")
            l.setFont(Fonts.label_sm())
            l.setStyleSheet(
                f"color: {Colors.TEXT_MUTED}; border: none;")
            gl.addWidget(l, row, col)
        vl.addWidget(grid)
        return card

    # ── Export actions ────────────────────────────────────────────────────────

    def _export_excel(self) -> None:
        """
        Export a premium executive Excel workbook.
        Big Four / Fortune 500 consulting quality.
        Presentation only — data is never modified.
        """
        import os
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        from core.database.db import (
            get_risks as _gr, get_treatments as _gt,
            get_organisation_scope, today as _today, audit,
        )

        try:
            import openpyxl
            from openpyxl.styles import (
                PatternFill, Font as XFont, Alignment,
                Border, Side, GradientFill,
            )
            from openpyxl.utils import get_column_letter
            from openpyxl.utils.cell import coordinate_from_string
            from openpyxl.chart import (
                DoughnutChart, BarChart, Reference, Series,
            )
            from openpyxl.chart.label import DataLabelList
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.worksheet.page import PageMargins
        except ImportError:
            QMessageBox.warning(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\n"
                "Run:  pip install openpyxl")
            return

        risks = [dict(r) for r in _gr()]
        if not risks:
            QMessageBox.information(self, "Export", "No risks to export.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Executive Workbook", "",
            "Excel Files (*.xlsx)")
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        # ══════════════════════════════════════════════════════════════════════
        # PALETTE — Enterprise dark-blue / slate palette
        # ══════════════════════════════════════════════════════════════════════
        P = {
            # Primary
            "navy":    "0B1F3A",   # darkest navy
            "blue":    "1B5FCF",   # primary blue
            "blue_lt": "2E75D4",   # lighter blue
            "slate":   "344563",   # slate header
            "mid":     "4A5568",   # mid grey text
            "dim":     "718096",   # dim text
            "border":  "CBD5E0",   # subtle border
            "row_alt": "F0F4F8",   # alternating row
            "white":   "FFFFFF",
            "bg":      "F7FAFC",   # page background
            # Severity
            "critical_bg":  "FDECEA",  "critical_fg":  "C62828",
            "high_bg":      "FFF3E0",  "high_fg":      "E65100",
            "medium_bg":    "FFFDE7",  "medium_fg":    "F57F17",
            "low_bg":       "E8F5E9",  "low_fg":       "2E7D32",
            # Status
            "open_bg":      "FFF3E0",  "open_fg":      "E65100",
            "inprog_bg":    "E3F2FD",  "inprog_fg":    "1565C0",
            "approved_bg":  "E8F5E9",  "approved_fg":  "2E7D32",
            "closed_bg":    "ECEFF1",  "closed_fg":    "546E7A",
            "draft_bg":     "F3E5F5",  "draft_fg":     "6A1B9A",
        }

        def F(c, ft="solid"):
            return PatternFill(ft, fgColor=c)

        def font(bold=False, color="1A2533", size=9, name="Calibri"):
            return XFont(name=name, bold=bold, color=color, size=size)

        def hdr_font(size=9, color="FFFFFF"):
            return XFont(name="Calibri", bold=True, color=color, size=size)

        def al(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def border(style="thin", color=None):
            c = color or P["border"]
            s = Side(style=style, color=c)
            return Border(left=s, right=s, top=s, bottom=s)

        def border_outer(color=None):
            c = color or P["blue"]
            s = Side(style="medium", color=c)
            return Border(left=s, right=s, top=s, bottom=s)

        def sev(score):
            s = int(score or 0)
            if s >= 15: return "CRITICAL", P["critical_fg"], P["critical_bg"]
            if s >= 10: return "HIGH",     P["high_fg"],     P["high_bg"]
            if s >= 5:  return "MEDIUM",   P["medium_fg"],   P["medium_bg"]
            return "LOW", P["low_fg"], P["low_bg"]

        STATUS_COLORS = {
            "Open":        (P["open_fg"],     P["open_bg"]),
            "In Progress": (P["inprog_fg"],   P["inprog_bg"]),
            "Approved":    (P["approved_fg"], P["approved_bg"]),
            "Completed":   (P["approved_fg"], P["approved_bg"]),
            "Verified":    (P["approved_fg"], P["approved_bg"]),
            "Closed":      (P["closed_fg"],   P["closed_bg"]),
            "Draft":       (P["draft_fg"],    P["draft_bg"]),
            "Mitigated":   (P["approved_fg"], P["approved_bg"]),
        }

        # ── Gather data ────────────────────────────────────────────────────────
        scope  = get_organisation_scope() or {}
        org    = scope.get("organisation_name") or self._org_name_from_settings()
        asmt   = scope.get("assessment_name", "Cyber Risk Assessment")
        today  = _today()

        total  = len(risks)
        crits  = sum(1 for r in risks if int(r.get("risk_score") or 0) >= 15)
        highs  = sum(1 for r in risks if 10 <= int(r.get("risk_score") or 0) <= 14)
        meds   = sum(1 for r in risks if 5  <= int(r.get("risk_score") or 0) <= 9)
        lows   = sum(1 for r in risks if int(r.get("risk_score") or 0) < 5)

        all_t = []
        for r in risks:
            for t in [dict(x) for x in _gt(r["id"])]:
                t["_risk_title"] = r.get("title", "")
                t["_risk_score"] = r.get("risk_score", 0)
                all_t.append(t)

        t_open  = sum(1 for t in all_t if t.get("status") in ("Draft","In Progress","Approved"))
        t_done  = sum(1 for t in all_t if t.get("status") in ("Completed","Verified"))

        # Overall posture
        avg_score = (sum(int(r.get("risk_score") or 0) for r in risks) / total) if total else 0
        if avg_score >= 15: posture = "HIGH RISK"
        elif avg_score >= 10: posture = "ELEVATED"
        elif avg_score >= 5: posture = "MODERATE"
        else: posture = "MANAGED"

        wb = openpyxl.Workbook()

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 1 — EXECUTIVE SUMMARY
        # ══════════════════════════════════════════════════════════════════════
        ws = wb.active
        ws.title = "Executive Summary"
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = P["navy"]

        # Column widths
        col_widths = {"A":3,"B":22,"C":22,"D":22,"E":22,"F":22,"G":22,"H":3}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        # Row heights
        for r in range(1, 60):
            ws.row_dimensions[r].height = 15

        # ── Header band ────────────────────────────────────────────────────────
        ws.row_dimensions[2].height = 14
        ws.row_dimensions[3].height = 36
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 22
        ws.row_dimensions[6].height = 18

        # Navy background rows 1-6
        for row in range(1, 7):
            for col in range(1, 9):
                ws.cell(row, col).fill = F(P["navy"])

        # Title
        ws.merge_cells("B3:G3")
        t = ws["B3"]
        t.value = f"CYBER RISK ASSESSMENT"
        t.font = XFont(name="Calibri", bold=True, color="FFFFFF", size=20)
        t.alignment = al("left", "center")

        # Subtitle
        ws.merge_cells("B4:G4")
        s = ws["B4"]
        s.value = f"{org}  ·  {asmt}"
        s.font = XFont(name="Calibri", bold=False, color="93C5FD", size=11)
        s.alignment = al("left", "center")

        # Meta row
        ws.merge_cells("B5:D5")
        m1 = ws["B5"]
        m1.value = f"Assessment Date: {today}"
        m1.font = XFont(name="Calibri", color="CBD5E0", size=9)
        m1.alignment = al("left", "center")

        ws.merge_cells("E5:G5")
        m2 = ws["E5"]
        m2.value = "Prepared by: RiskCore GRC Platform v1.5"
        m2.font = XFont(name="Calibri", color="CBD5E0", size=9)
        m2.alignment = al("right", "center")

        # Blue accent line
        ws.row_dimensions[7].height = 4
        for col in range(1, 9):
            ws.cell(7, col).fill = F(P["blue"])

        # ── KPI Cards (row 9-16) ───────────────────────────────────────────────
        ws.row_dimensions[8].height = 12
        ws.row_dimensions[9].height = 14
        ws.row_dimensions[10].height = 28
        ws.row_dimensions[11].height = 16
        ws.row_dimensions[12].height = 14

        kpi_data = [
            ("B9:B12",  "B",  "TOTAL\nRISKS",   total,   P["blue"],       P["white"]),
            ("C9:C12",  "C",  "CRITICAL",        crits,   P["critical_fg"],P["critical_bg"]),
            ("D9:D12",  "D",  "HIGH",            highs,   P["high_fg"],    P["high_bg"]),
            ("E9:E12",  "E",  "MEDIUM",          meds,    P["medium_fg"],  P["medium_bg"]),
            ("F9:F12",  "F",  "LOW",             lows,    P["low_fg"],     P["low_bg"]),
            ("G9:G12",  "G",  "OPEN\nTREATMENTS",t_open, P["inprog_fg"],  P["inprog_bg"]),
        ]

        for merge_rng, col, lbl, val, fg, bg in kpi_data:
            # Label row
            lbl_cell = ws[f"{col}9"]
            lbl_cell.value = lbl
            lbl_cell.font = XFont(name="Calibri", bold=True, color=fg, size=8)
            lbl_cell.fill = F(bg)
            lbl_cell.alignment = al("center", "center", wrap=True)
            # Value row
            val_cell = ws[f"{col}10"]
            val_cell.value = val
            val_cell.font = XFont(name="Calibri", bold=True, color=fg, size=22)
            val_cell.fill = F(bg)
            val_cell.alignment = al("center", "center")
            # Bottom label row
            bot_cell = ws[f"{col}11"]
            bot_cell.value = "risks" if "RISK" in lbl or lbl in ("CRITICAL","HIGH","MEDIUM","LOW") else "items"
            bot_cell.font = XFont(name="Calibri", color=fg, size=8)
            bot_cell.fill = F(bg)
            bot_cell.alignment = al("center", "top")
            # Borders
            for rn in [9, 10, 11]:
                ws[f"{col}{rn}"].border = border_outer(fg)

        # Overall posture card
        ws.row_dimensions[13].height = 8
        ws.merge_cells("B14:G14")
        pos_lbl = ws["B14"]
        pos_lbl.value = "OVERALL SECURITY POSTURE"
        pos_lbl.font = XFont(name="Calibri", bold=True, color=P["dim"], size=8)
        pos_lbl.alignment = al("center", "center")

        ws.row_dimensions[14].height = 14
        ws.merge_cells("B15:G16")
        pos_val = ws["B15"]
        pos_val.value = posture
        pos_c = P["critical_fg"] if posture == "HIGH RISK" else \
                P["high_fg"] if posture == "ELEVATED" else \
                P["medium_fg"] if posture == "MODERATE" else P["low_fg"]
        pos_bg = P["critical_bg"] if posture == "HIGH RISK" else \
                 P["high_bg"] if posture == "ELEVATED" else \
                 P["medium_bg"] if posture == "MODERATE" else P["low_bg"]
        pos_val.font = XFont(name="Calibri", bold=True, color=pos_c, size=18)
        pos_val.fill = F(pos_bg)
        pos_val.alignment = al("center", "center")
        pos_val.border = border_outer(pos_c)
        ws.row_dimensions[15].height = 26
        ws.row_dimensions[16].height = 10

        # ── Summary table (row 18+) ────────────────────────────────────────────
        ws.row_dimensions[18].height = 14
        ws["B18"].value = "RISK SUMMARY"
        ws["B18"].font = XFont(name="Calibri", bold=True, color=P["navy"], size=10)
        ws["B18"].border = Border(bottom=Side(style="medium", color=P["blue"]))

        ws.row_dimensions[19].height = 20
        sum_headers = ["Severity", "Count", "% of Total", "Avg Score", "Status"]
        for ci, h in enumerate(sum_headers, 2):
            c = ws.cell(19, ci, h)
            c.font = hdr_font(size=9)
            c.fill = F(P["slate"])
            c.alignment = al("center", "center")
            c.border = border(color=P["slate"])

        sum_rows = [
            ("Critical (≥15)", crits,  P["critical_fg"]),
            ("High (10–14)",   highs,  P["high_fg"]),
            ("Medium (5–9)",   meds,   P["medium_fg"]),
            ("Low (<5)",       lows,   P["low_fg"]),
            ("Total",          total,  P["blue"]),
        ]
        for i, (lbl, cnt, col) in enumerate(sum_rows, start=20):
            ws.row_dimensions[i].height = 18
            alt = i % 2 == 0
            bg = P["row_alt"] if alt else P["white"]
            is_total = lbl == "Total"

            c1 = ws.cell(i, 2, lbl)
            c1.font = XFont(name="Calibri", bold=is_total, color=col, size=9)
            c1.fill = F(bg); c1.alignment = al("left","center")
            c1.border = border()

            c2 = ws.cell(i, 3, cnt)
            c2.font = XFont(name="Calibri", bold=is_total, color=col, size=9)
            c2.fill = F(bg); c2.alignment = al("center","center")
            c2.border = border()

            pct = f"{cnt/total*100:.0f}%" if total > 0 and not is_total else ""
            c3 = ws.cell(i, 4, pct)
            c3.font = font(bold=is_total, size=9)
            c3.fill = F(bg); c3.alignment = al("center","center")
            c3.border = border()

            # Avg score per severity
            if not is_total:
                sev_min = {"Critical (≥15)": 15, "High (10–14)": 10,
                           "Medium (5–9)": 5, "Low (<5)": 0}[lbl]
                sev_max = {"Critical (≥15)": 99, "High (10–14)": 14,
                           "Medium (5–9)": 9, "Low (<5)": 4}[lbl]
                sev_risks = [r for r in risks
                             if sev_min <= int(r.get("risk_score") or 0) <= sev_max]
                avg = (sum(int(r.get("risk_score") or 0) for r in sev_risks)
                       / len(sev_risks)) if sev_risks else 0
                c4 = ws.cell(i, 5, f"{avg:.1f}" if avg else "—")
            else:
                c4 = ws.cell(i, 5, f"{avg_score:.1f}")
            c4.font = font(bold=is_total, size=9)
            c4.fill = F(bg); c4.alignment = al("center","center")
            c4.border = border()

            c5 = ws.cell(i, 6, "")
            c5.fill = F(bg); c5.border = border()

        # ── NIST Coverage table ────────────────────────────────────────────────
        ws.row_dimensions[26].height = 14
        ws["B26"].value = "NIST CSF 2.0 COVERAGE"
        ws["B26"].font = XFont(name="Calibri", bold=True, color=P["navy"], size=10)
        ws["B26"].border = Border(bottom=Side(style="medium", color=P["blue"]))

        ws.row_dimensions[27].height = 20
        for ci, h in enumerate(["Function", "Risks", "Coverage"], 2):
            c = ws.cell(27, ci, h)
            c.font = hdr_font(size=9); c.fill = F(P["slate"])
            c.alignment = al("center","center"); c.border = border(color=P["slate"])

        nist_fns = ["Govern","Identify","Protect","Detect","Respond","Recover"]
        nist_colors = {
            "Govern": "5C6BC0", "Identify": "1565C0",
            "Protect": "00838F", "Detect": "2E7D32",
            "Respond": "E65100", "Recover": "6A1B9A",
        }
        for i, fn in enumerate(nist_fns, start=28):
            ws.row_dimensions[i].height = 18
            cnt_fn = sum(1 for r in risks if r.get("nist_function") == fn)
            alt = i % 2 == 0
            bg = P["row_alt"] if alt else P["white"]
            col = nist_colors.get(fn, P["blue"])

            c1 = ws.cell(i, 2, fn)
            c1.font = XFont(name="Calibri", bold=True, color=col, size=9)
            c1.fill = F(bg); c1.alignment = al("left","center"); c1.border = border()

            c2 = ws.cell(i, 3, cnt_fn)
            c2.font = font(size=9); c2.fill = F(bg)
            c2.alignment = al("center","center"); c2.border = border()

            pct_str = f"{cnt_fn/total*100:.0f}%" if total > 0 else "0%"
            c3 = ws.cell(i, 4, pct_str)
            c3.font = font(size=9); c3.fill = F(bg)
            c3.alignment = al("center","center"); c3.border = border()

        # ── Footer ─────────────────────────────────────────────────────────────
        ws.row_dimensions[36].height = 20
        ws.merge_cells("B36:G36")
        ft = ws["B36"]
        ft.value = (f"CONFIDENTIAL  ·  {org}  ·  Generated by RiskCore GRC Platform v1.5  ·  {today}")
        ft.font = XFont(name="Calibri", color=P["dim"], size=8, italic=True)
        ft.fill = F(P["bg"]); ft.alignment = al("center","center")

        # Print settings
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        ws.oddHeader.center.text = f"&B{org} — Cyber Risk Assessment"
        ws.oddFooter.left.text = "CONFIDENTIAL"
        ws.oddFooter.center.text = "RiskCore GRC Platform v1.5"
        ws.oddFooter.right.text = "Page &P of &N"

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 2 — RISK REGISTER
        # ══════════════════════════════════════════════════════════════════════
        ws_r = wb.create_sheet("Risk Register")
        ws_r.sheet_view.showGridLines = False
        ws_r.freeze_panes = "C2"
        ws_r.sheet_properties.tabColor = P["navy"]

        r_cols = [
            ("ID",           "id",                 6,  "center"),
            ("Title",        "title",              36, "left"),
            ("Severity",     "_sev",               11, "center"),
            ("Score",        "risk_score",         7,  "center"),
            ("Likelihood",   "likelihood",         10, "center"),
            ("Impact",       "impact",             8,  "center"),
            ("Residual",     "residual_score",     9,  "center"),
            ("Status",       "status",             13, "center"),
            ("Priority",     "priority",           12, "center"),
            ("Owner",        "owner",              18, "left"),
            ("NIST Function","nist_function",      14, "left"),
            ("NIST Category","nist_category",      20, "left"),
            ("ISO Domain",   "iso_domain",         22, "left"),
            ("ISO Control",  "iso_control",        10, "center"),
            ("MITRE Tactic", "mitre_tactic",       18, "left"),
            ("CIS Control",  "cis_control",        10, "center"),
            ("CIA",          "cia_component",      12, "center"),
            ("Review Date",  "review_date",        12, "center"),
            ("Source",       "source",             10, "center"),
            ("Description",  "description",        40, "left"),
            ("Mitigation",   "mitigation",         35, "left"),
            ("Controls",     "existing_controls",  28, "left"),
        ]

        # Header row
        ws_r.row_dimensions[1].height = 26
        for ci, (hdr, _, w, _al) in enumerate(r_cols, 1):
            c = ws_r.cell(1, ci, hdr)
            c.font = hdr_font(size=9)
            c.fill = F(P["navy"])
            c.alignment = al("center", "center", wrap=True)
            c.border = border(color=P["navy"])
            ws_r.column_dimensions[get_column_letter(ci)].width = w

        # Data rows
        for ri, r in enumerate(risks, start=2):
            slbl, sfg, sbg = sev(r.get("risk_score", 0))
            alt = ri % 2 == 0
            row_bg = P["row_alt"] if alt else P["white"]
            ws_r.row_dimensions[ri].height = 20

            for ci, (_, key, _, cell_al) in enumerate(r_cols, 1):
                if key == "_sev":
                    val = slbl
                else:
                    val = r.get(key, "") or ""
                    if isinstance(val, float) and val == int(val):
                        val = int(val)

                c = ws_r.cell(ri, ci, val)
                c.border = border()
                c.alignment = al(cell_al, "center",
                                 wrap=True if cell_al == "left" else False)

                if key == "_sev":
                    c.font = XFont(name="Calibri", bold=True, color=sfg, size=9)
                    c.fill = F(sbg)
                elif key == "risk_score":
                    c.font = XFont(name="Calibri", bold=True, color=sfg, size=9)
                    c.fill = F(sbg)
                elif key == "status":
                    st = str(val)
                    st_fg, st_bg = STATUS_COLORS.get(st, (P["mid"], P["white"]))
                    c.font = XFont(name="Calibri", bold=True, color=st_fg, size=9)
                    c.fill = F(st_bg)
                elif key == "residual_score":
                    if val and int(val) > 0:
                        _, rfg, rbg = sev(int(val))
                        c.font = XFont(name="Calibri", bold=True, color=rfg, size=9)
                        c.fill = F(rbg)
                    else:
                        c.font = font(size=9); c.fill = F(row_bg)
                elif key == "priority":
                    p_colors = {
                        "Immediate":   (P["critical_fg"], P["critical_bg"]),
                        "Short-term":  (P["high_fg"],     P["high_bg"]),
                        "Medium-term": (P["medium_fg"],   P["medium_bg"]),
                        "Long-term":   (P["low_fg"],      P["low_bg"]),
                    }
                    pfg, pbg = p_colors.get(str(val), (P["mid"], row_bg))
                    c.font = XFont(name="Calibri", bold=True, color=pfg, size=9)
                    c.fill = F(pbg)
                else:
                    c.font = font(size=9)
                    c.fill = F(row_bg)

        # Excel Table
        if len(risks) > 0:
            last_col = get_column_letter(len(r_cols))
            last_row = len(risks) + 1
            tbl = Table(
                displayName="RiskRegister",
                ref=f"A1:{last_col}{last_row}")
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=False, showColumnStripes=False)
            ws_r.add_table(tbl)

        # Print settings
        ws_r.page_setup.orientation = "landscape"
        ws_r.page_setup.fitToPage = True
        ws_r.page_setup.fitToWidth = 1
        ws_r.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        ws_r.print_title_rows = "1:1"
        ws_r.oddHeader.left.text = f"&B{org} — Risk Register"
        ws_r.oddHeader.right.text = "&B CONFIDENTIAL"
        ws_r.oddFooter.left.text = "RiskCore GRC Platform v1.5"
        ws_r.oddFooter.right.text = "Page &P of &N"

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 3 — TREATMENTS
        # ══════════════════════════════════════════════════════════════════════
        ws_t = wb.create_sheet("Treatments")
        ws_t.sheet_view.showGridLines = False
        ws_t.freeze_panes = "B2"
        ws_t.sheet_properties.tabColor = P["blue"]

        t_cols = [
            ("Risk ID",        "risk_id",              7,  "center"),
            ("Risk Title",     "_risk_title",           32, "left"),
            ("Risk Score",     "_risk_score",           9,  "center"),
            ("Treatment",      "title",                 28, "left"),
            ("Strategy",       "strategy",              12, "center"),
            ("Status",         "status",                14, "center"),
            ("Owner",          "owner",                 18, "left"),
            ("Target Date",    "target_date",           13, "center"),
            ("Residual Target","residual_score_target", 12, "center"),
            ("Cost Estimate",  "cost_estimate",         14, "center"),
            ("Description",    "description",           38, "left"),
        ]

        STRATEGY_COLORS = {
            "Mitigate":  (P["inprog_fg"],   P["inprog_bg"]),
            "Accept":    (P["low_fg"],       P["low_bg"]),
            "Transfer":  (P["medium_fg"],    P["medium_bg"]),
            "Avoid":     (P["critical_fg"],  P["critical_bg"]),
        }

        ws_t.row_dimensions[1].height = 26
        for ci, (hdr, _, w, _al) in enumerate(t_cols, 1):
            c = ws_t.cell(1, ci, hdr)
            c.font = hdr_font(size=9)
            c.fill = F(P["blue"])
            c.alignment = al("center", "center", wrap=True)
            c.border = border(color=P["blue"])
            ws_t.column_dimensions[get_column_letter(ci)].width = w

        for ti, t in enumerate(all_t, start=2):
            alt = ti % 2 == 0
            row_bg = P["row_alt"] if alt else P["white"]
            ws_t.row_dimensions[ti].height = 20

            for ci, (_, key, _, cell_al) in enumerate(t_cols, 1):
                val = t.get(key, "") or ""
                if isinstance(val, float) and val == int(val):
                    val = int(val)

                c = ws_t.cell(ti, ci, val)
                c.border = border()
                c.alignment = al(cell_al, "center",
                                 wrap=True if cell_al == "left" else False)

                if key == "_risk_score":
                    _, rfg, rbg = sev(int(val) if val else 0)
                    c.font = XFont(name="Calibri", bold=True, color=rfg, size=9)
                    c.fill = F(rbg)
                elif key == "status":
                    st = str(val)
                    st_fg, st_bg = STATUS_COLORS.get(st, (P["mid"], row_bg))
                    c.font = XFont(name="Calibri", bold=True, color=st_fg, size=9)
                    c.fill = F(st_bg)
                elif key == "strategy":
                    strat_fg, strat_bg = STRATEGY_COLORS.get(str(val), (P["mid"], row_bg))
                    c.font = XFont(name="Calibri", bold=True, color=strat_fg, size=9)
                    c.fill = F(strat_bg)
                elif key == "cost_estimate":
                    if val and str(val).replace(".", "").isdigit():
                        c.number_format = '"£"#,##0'
                    c.font = font(size=9); c.fill = F(row_bg)
                elif key == "residual_score_target":
                    if val and int(val) > 0:
                        _, rfg, rbg = sev(int(val))
                        c.font = XFont(name="Calibri", bold=True, color=rfg, size=9)
                        c.fill = F(rbg)
                    else:
                        c.font = font(size=9); c.fill = F(row_bg)
                else:
                    c.font = font(size=9)
                    c.fill = F(row_bg)

        # Excel Table
        if all_t:
            last_col_t = get_column_letter(len(t_cols))
            last_row_t = len(all_t) + 1
            tbl_t = Table(
                displayName="TreatmentPlan",
                ref=f"A1:{last_col_t}{last_row_t}")
            tbl_t.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=False, showColumnStripes=False)
            ws_t.add_table(tbl_t)

        # Print settings
        ws_t.page_setup.orientation = "landscape"
        ws_t.page_setup.fitToPage = True
        ws_t.page_setup.fitToWidth = 1
        ws_t.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        ws_t.print_title_rows = "1:1"
        ws_t.oddHeader.left.text = f"&B{org} — Treatment Roadmap"
        ws_t.oddHeader.right.text = "&B CONFIDENTIAL"
        ws_t.oddFooter.left.text = "RiskCore GRC Platform v1.5"
        ws_t.oddFooter.right.text = "Page &P of &N"


        # ══════════════════════════════════════════════════════════════════════
        # SHEET 4 — CHARTS & ANALYTICS
        # ══════════════════════════════════════════════════════════════════════
        from openpyxl.chart import BarChart, DoughnutChart, PieChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.chart.label import DataLabelList

        ws_c = wb.create_sheet("Charts & Analytics")
        ws_c.sheet_view.showGridLines = False
        ws_c.sheet_properties.tabColor = P["slate"]

        # ── Hidden data tables for charts ─────────────────────────────────────
        # Severity data (col A-B, rows 2-6)
        ws_c["A1"].value = "Severity"; ws_c["B1"].value = "Count"
        for i, (lbl, cnt) in enumerate([
            ("Critical", crits), ("High", highs),
            ("Medium", meds), ("Low", lows)
        ], start=2):
            ws_c.cell(i, 1, lbl); ws_c.cell(i, 2, cnt)

        # Treatment status data (col D-E, rows 2-8)
        ws_c["D1"].value = "Status"; ws_c["E1"].value = "Count"
        t_statuses = {}
        for t in all_t:
            s = t.get("status", "Draft")
            t_statuses[s] = t_statuses.get(s, 0) + 1
        for i, (s, cnt) in enumerate(t_statuses.items(), start=2):
            ws_c.cell(i, 4, s); ws_c.cell(i, 5, cnt)
        t_last = len(t_statuses) + 1

        # NIST function data (col G-H, rows 2-8)
        ws_c["G1"].value = "NIST Function"; ws_c["H1"].value = "Risks"
        nist_fns = ["Govern","Identify","Protect","Detect","Respond","Recover"]
        for i, fn in enumerate(nist_fns, start=2):
            cnt_fn = sum(1 for r in risks if r.get("nist_function") == fn)
            ws_c.cell(i, 7, fn); ws_c.cell(i, 8, cnt_fn)

        # CIA data (col J-K, rows 2-6)
        ws_c["J1"].value = "CIA Component"; ws_c["K1"].value = "Risks"
        cia_counts = {}
        for r in risks:
            c = r.get("cia_component","Unknown")
            cia_counts[c] = cia_counts.get(c, 0) + 1
        for i, (c, cnt) in enumerate(cia_counts.items(), start=2):
            ws_c.cell(i, 10, c); ws_c.cell(i, 11, cnt)
        cia_last = len(cia_counts) + 1

        # Category data (col M-N)
        ws_c["M1"].value = "Category"; ws_c["N1"].value = "Risks"
        cat_counts = {}
        for r in risks:
            cat = r.get("category","Other")
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
        for i, (cat, cnt) in enumerate(cat_counts.items(), start=2):
            ws_c.cell(i, 13, cat); ws_c.cell(i, 14, cnt)
        cat_last = len(cat_counts) + 1

        # ── Chart 1: Severity Distribution (Doughnut) ─────────────────────────
        sev_chart = DoughnutChart()
        sev_chart.title = "Risk Severity Distribution"
        sev_chart.style = 10
        sev_chart.holeSize = 50
        sev_labels = Reference(ws_c, min_col=1, min_row=2, max_row=5)
        sev_data   = Reference(ws_c, min_col=2, min_row=1, max_row=5)
        sev_chart.add_data(sev_data, titles_from_data=True)
        sev_chart.set_categories(sev_labels)
        sev_chart.dataLabels = DataLabelList()
        sev_chart.dataLabels.showPercent = True
        sev_chart.dataLabels.showCatName = False
        sev_chart.dataLabels.showVal = False
        sev_chart.width  = 14
        sev_chart.height = 10
        # Colour the slices
        from openpyxl.chart.data_source import NumDataSource, NumRef
        from openpyxl.drawing.fill import ColorChoice
        sev_chart.series[0].graphicalProperties.solidFill = P["critical_fg"]
        ws_c.add_chart(sev_chart, "A8")

        # ── Chart 2: NIST CSF Coverage (Horizontal Bar) ───────────────────────
        nist_chart = BarChart()
        nist_chart.type    = "bar"
        nist_chart.barDir  = "bar"
        nist_chart.style   = 10
        nist_chart.title   = "Risk Distribution by NIST CSF 2.0 Function"
        nist_chart.y_axis.title = "NIST Function"
        nist_chart.x_axis.title = "Number of Risks"
        nist_chart.y_axis.majorGridlines = None
        nist_labels = Reference(ws_c, min_col=7, min_row=2, max_row=7)
        nist_data   = Reference(ws_c, min_col=8, min_row=1, max_row=7)
        nist_chart.add_data(nist_data, titles_from_data=True)
        nist_chart.set_categories(nist_labels)
        nist_chart.width  = 18
        nist_chart.height = 12
        ws_c.add_chart(nist_chart, "H8")

        # ── Chart 3: Treatment Status (Doughnut) ──────────────────────────────
        if t_statuses:
            treat_chart = DoughnutChart()
            treat_chart.title    = "Treatment Status"
            treat_chart.style    = 10
            treat_chart.holeSize = 50
            t_labels = Reference(ws_c, min_col=4, min_row=2, max_row=t_last)
            t_data   = Reference(ws_c, min_col=5, min_row=1, max_row=t_last)
            treat_chart.add_data(t_data, titles_from_data=True)
            treat_chart.set_categories(t_labels)
            treat_chart.dataLabels = DataLabelList()
            treat_chart.dataLabels.showPercent = True
            treat_chart.width  = 14
            treat_chart.height = 10
            ws_c.add_chart(treat_chart, "A30")

        # ── Chart 4: CIA Triad Distribution (Bar) ─────────────────────────────
        if cia_counts:
            cia_chart = BarChart()
            cia_chart.type   = "col"
            cia_chart.style  = 10
            cia_chart.title  = "Risks by CIA Component"
            cia_chart.y_axis.title = "Risks"
            cia_chart.y_axis.majorGridlines = None
            cia_l = Reference(ws_c, min_col=10, min_row=2, max_row=cia_last)
            cia_d = Reference(ws_c, min_col=11, min_row=1, max_row=cia_last)
            cia_chart.add_data(cia_d, titles_from_data=True)
            cia_chart.set_categories(cia_l)
            cia_chart.width  = 14
            cia_chart.height = 10
            ws_c.add_chart(cia_chart, "H30")

        # ── Chart 5: Risk Category Breakdown (Horizontal Bar) ─────────────────
        if cat_counts:
            cat_chart = BarChart()
            cat_chart.type   = "bar"
            cat_chart.barDir = "bar"
            cat_chart.style  = 10
            cat_chart.title  = "Risks by Category"
            cat_chart.y_axis.majorGridlines = None
            cat_l = Reference(ws_c, min_col=13, min_row=2, max_row=cat_last)
            cat_d = Reference(ws_c, min_col=14, min_row=1, max_row=cat_last)
            cat_chart.add_data(cat_d, titles_from_data=True)
            cat_chart.set_categories(cat_l)
            cat_chart.width  = 18
            cat_chart.height = 10
            ws_c.add_chart(cat_chart, "A50")

        # Print settings
        ws_c.page_setup.orientation = "landscape"
        ws_c.page_setup.fitToPage  = True
        ws_c.page_setup.fitToWidth = 1
        ws_c.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        ws_c.oddHeader.center.text = f"&B{org} — Risk Analytics"
        ws_c.oddFooter.right.text  = "Page &P of &N"

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 5 — PIVOT-STYLE ANALYSIS
        # ══════════════════════════════════════════════════════════════════════
        ws_p = wb.create_sheet("Risk Analysis")
        ws_p.sheet_view.showGridLines = False
        ws_p.sheet_properties.tabColor = P["blue_lt"]

        # Column widths
        for col, w in {"A":3,"B":24,"C":12,"D":12,"E":12,"F":12,"G":12,"H":12,"I":3}.items():
            ws_p.column_dimensions[col].width = w

        def ph(row, col, value, bg=None, bold=True, size=9, color="FFFFFF"):
            c = ws_p.cell(row, col, value)
            c.font = XFont(name="Calibri", bold=bold, color=color, size=size)
            if bg: c.fill = F(bg)
            c.alignment = al("center", "center")
            c.border = border()
            return c

        def pd(row, col, value, bg=None, bold=False, align="center", color="1A2533"):
            c = ws_p.cell(row, col, value)
            c.font = XFont(name="Calibri", bold=bold, color=color, size=9)
            if bg: c.fill = F(bg)
            c.alignment = al(align, "center")
            c.border = border()
            return c

        # Sheet title
        ws_p.row_dimensions[2].height = 30
        ws_p.merge_cells("B2:H2")
        t = ws_p["B2"]
        t.value = "RISK ANALYSIS — PIVOT VIEW"
        t.font = XFont(name="Calibri", bold=True, color="FFFFFF", size=14)
        t.fill = F(P["navy"]); t.alignment = al("center","center")

        ws_p.row_dimensions[3].height = 16
        ws_p.merge_cells("B3:H3")
        s = ws_p["B3"]
        s.value = f"{org}  ·  {asmt}  ·  {today}"
        s.font = XFont(name="Calibri", color=P["dim"], size=9)
        s.fill = F(P["bg"]); s.alignment = al("center","center")

        # ── Table 1: Risks by NIST Function × Severity ────────────────────────
        ws_p.row_dimensions[5].height = 18
        ws_p.merge_cells("B5:H5")
        ws_p["B5"].value = "RISKS BY NIST CSF 2.0 FUNCTION × SEVERITY"
        ws_p["B5"].font = XFont(name="Calibri", bold=True, color=P["navy"], size=10)
        ws_p["B5"].border = Border(bottom=Side(style="medium", color=P["blue"]))

        ws_p.row_dimensions[6].height = 20
        for ci, hdr in enumerate(["Function","Critical","High","Medium","Low","Total","Avg Score"],2):
            ph(6, ci, hdr, bg=P["navy"])

        nist_rows = []
        for fn in nist_fns:
            fn_risks = [r for r in risks if r.get("nist_function") == fn]
            fn_crits = sum(1 for r in fn_risks if int(r.get("risk_score") or 0) >= 15)
            fn_highs = sum(1 for r in fn_risks if 10 <= int(r.get("risk_score") or 0) <= 14)
            fn_meds  = sum(1 for r in fn_risks if 5  <= int(r.get("risk_score") or 0) <= 9)
            fn_lows  = sum(1 for r in fn_risks if int(r.get("risk_score") or 0) < 5)
            fn_total = len(fn_risks)
            fn_avg   = (sum(int(r.get("risk_score") or 0) for r in fn_risks) / fn_total
                        if fn_total else 0)
            nist_rows.append((fn, fn_crits, fn_highs, fn_meds, fn_lows, fn_total, fn_avg))

        for i, row_data in enumerate(nist_rows, start=7):
            ws_p.row_dimensions[i].height = 18
            alt = i % 2 == 0; bg = P["row_alt"] if alt else P["white"]
            fn, cr, hi, me, lo, tot, avg = row_data
            fn_col = nist_colors.get(fn, P["blue"])
            pd(i, 2, fn, bg=bg, align="left", color=fn_col, bold=True)
            pd(i, 3, cr or "—", bg=P["critical_bg"] if cr else bg,
               color=P["critical_fg"] if cr else P["dim"])
            pd(i, 4, hi or "—", bg=P["high_bg"] if hi else bg,
               color=P["high_fg"] if hi else P["dim"])
            pd(i, 5, me or "—", bg=P["medium_bg"] if me else bg,
               color=P["medium_fg"] if me else P["dim"])
            pd(i, 6, lo or "—", bg=P["low_bg"] if lo else bg,
               color=P["low_fg"] if lo else P["dim"])
            pd(i, 7, tot, bg=bg, bold=True, color=P["blue"])
            pd(i, 8, f"{avg:.1f}" if avg else "—", bg=bg)

        # Totals row
        tr = len(nist_rows) + 7
        ws_p.row_dimensions[tr].height = 20
        for ci, val in enumerate([
            "TOTAL", crits, highs, meds, lows, total,
            f"{avg_score:.1f}"
        ], 2):
            c = ws_p.cell(tr, ci, val)
            c.font = XFont(name="Calibri", bold=True, color=P["white"], size=9)
            c.fill = F(P["slate"]); c.alignment = al("center","center"); c.border = border()

        # ── Table 2: Treatment Summary by Strategy ─────────────────────────────
        start_row = tr + 3
        ws_p.row_dimensions[start_row].height = 18
        ws_p.merge_cells(f"B{start_row}:H{start_row}")
        ws_p[f"B{start_row}"].value = "TREATMENT PLAN SUMMARY BY STRATEGY"
        ws_p[f"B{start_row}"].font = XFont(name="Calibri", bold=True, color=P["navy"], size=10)
        ws_p[f"B{start_row}"].border = Border(bottom=Side(style="medium", color=P["blue"]))

        hdr_row = start_row + 1
        ws_p.row_dimensions[hdr_row].height = 20
        for ci, hdr in enumerate(["Strategy","Total","Approved","In Progress","Draft","Completed"],2):
            ph(hdr_row, ci, hdr, bg=P["blue"])

        strategies = ["Mitigate","Accept","Transfer","Avoid"]
        for i, strat in enumerate(strategies, start=hdr_row+1):
            ws_p.row_dimensions[i].height = 18
            alt = i % 2 == 0; bg = P["row_alt"] if alt else P["white"]
            strat_t = [t for t in all_t if t.get("strategy") == strat]
            total_s   = len(strat_t)
            approved  = sum(1 for t in strat_t if t.get("status") == "Approved")
            in_prog   = sum(1 for t in strat_t if t.get("status") == "In Progress")
            draft     = sum(1 for t in strat_t if t.get("status") == "Draft")
            completed = sum(1 for t in strat_t if t.get("status") in ("Completed","Verified"))
            strat_fg, strat_bg = STRATEGY_COLORS.get(strat, (P["mid"], bg))
            pd(i, 2, strat, bg=strat_bg, align="left", color=strat_fg, bold=True)
            pd(i, 3, total_s or "—", bg=bg, bold=True)
            pd(i, 4, approved  or "—", bg=P["approved_bg"] if approved else bg,
               color=P["approved_fg"] if approved else P["dim"])
            pd(i, 5, in_prog   or "—", bg=P["inprog_bg"] if in_prog else bg,
               color=P["inprog_fg"] if in_prog else P["dim"])
            pd(i, 6, draft     or "—", bg=P["draft_bg"] if draft else bg,
               color=P["draft_fg"] if draft else P["dim"])
            pd(i, 7, completed or "—", bg=P["approved_bg"] if completed else bg,
               color=P["approved_fg"] if completed else P["dim"])

        # ── Table 3: Top 5 Critical Risks ──────────────────────────────────────
        top_start = hdr_row + len(strategies) + 3
        ws_p.row_dimensions[top_start].height = 18
        ws_p.merge_cells(f"B{top_start}:H{top_start}")
        ws_p[f"B{top_start}"].value = "TOP RISKS BY SCORE — IMMEDIATE ATTENTION REQUIRED"
        ws_p[f"B{top_start}"].font = XFont(name="Calibri", bold=True, color=P["navy"], size=10)
        ws_p[f"B{top_start}"].border = Border(bottom=Side(style="medium", color=P["critical_fg"]))

        top_hdr = top_start + 1
        ws_p.row_dimensions[top_hdr].height = 20
        for ci, hdr in enumerate(["Risk Title","Score","Severity","Owner","Status","NIST Function"],2):
            ph(top_hdr, ci, hdr, bg=P["critical_fg"])

        top_risks_sorted = sorted(
            risks, key=lambda r: int(r.get("risk_score") or 0), reverse=True)[:8]

        for i, r in enumerate(top_risks_sorted, start=top_hdr+1):
            ws_p.row_dimensions[i].height = 18
            alt = i % 2 == 0; bg = P["row_alt"] if alt else P["white"]
            sc = int(r.get("risk_score") or 0)
            slbl, sfg, sbg = sev(sc)

            title_cell = ws_p.cell(i, 2, r.get("title","")[:60])
            title_cell.font = XFont(name="Calibri", size=9, color="1A2533")
            title_cell.fill = F(bg); title_cell.alignment = al("left","center",wrap=True)
            title_cell.border = border()

            score_cell = ws_p.cell(i, 3, sc)
            score_cell.font = XFont(name="Calibri", bold=True, color=sfg, size=10)
            score_cell.fill = F(sbg); score_cell.alignment = al("center","center")
            score_cell.border = border()

            pd(i, 4, slbl, bg=sbg, bold=True, color=sfg)
            pd(i, 5, r.get("owner","—"), bg=bg, align="left")
            pd(i, 6, r.get("status","—"),
               bg=STATUS_COLORS.get(r.get("status",""),("",""))[1] or bg,
               color=STATUS_COLORS.get(r.get("status",""),("",""))[0] or P["mid"],
               bold=True)
            pd(i, 7, r.get("nist_function","—"), bg=bg,
               color=nist_colors.get(r.get("nist_function",""), P["blue"]), bold=True)

        # Print settings
        ws_p.page_setup.orientation = "landscape"
        ws_p.page_setup.fitToPage  = True
        ws_p.page_setup.fitToWidth = 1
        ws_p.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        ws_p.print_title_rows = "2:3"
        ws_p.oddHeader.center.text = f"&B{org} — Risk Analysis"
        ws_p.oddFooter.left.text   = "CONFIDENTIAL"
        ws_p.oddFooter.right.text  = "Page &P of &N"

        # ── Workbook properties ────────────────────────────────────────────────
        wb.properties.title    = f"{org} — Cyber Risk Assessment"
        wb.properties.subject  = asmt
        wb.properties.creator  = "RiskCore GRC Platform v1.5"
        wb.properties.company  = org
        wb.properties.keywords = "GRC, Risk Assessment, NIST CSF, ISO 27001"

        # ── Save ───────────────────────────────────────────────────────────────
        try:
            wb.save(path)
            audit("EXPORT_CSV",
                  detail=f"Excel: {len(risks)} risks, "
                         f"{len(all_t)} treatments → {os.path.basename(path)}")
            self._set_status(
                f"✅  Excel exported: {os.path.basename(path)}",
                Colors.SUCCESS_LT)
        except Exception as e:
            self._set_status(f"⚠  Excel error: {e}", Colors.CRITICAL)


    def _set_status(self, text: str,
                    color: str = Colors.TEXT_MUTED) -> None:
        self._status_lbl.setText(text)
        self._status_lbl.setStyleSheet(
            f"color: {color}; border: none;")

    def _export_pdf(self) -> None:
        company = self._company_input.text().strip() or                   "Your Organisation"
        clf     = self._clf_combo.currentText()
        # Always query fresh — never trust stale cache
        from core.database.db import get_risks as _gr
        self._risks_cache = [dict(r) for r in _gr()]
        risks   = self._risks_cache
        if not risks:
            QMessageBox.information(
                self, "No Risks",
                "Add risks to the register before exporting.")
            return

        # Prepare risk dicts
        risks_out = []
        for r in risks:
            rd = dict(r)
            if not rd.get("inherent_score"):
                rd["inherent_score"] = int(
                    rd.get("risk_score") or 0)
            if not rd.get("residual_score"):
                rd["residual_score"] = max(
                    1, rd["inherent_score"] - 2)
            risks_out.append(rd)

        # Use cached analysis if available — avoids blocking API call
        # If no cache, use a static data-only analysis (no API call)
        # Full AI analysis happens in AI Analysis workspace, not here
        if self._cached_analysis:
            analysis = self._cached_analysis
        else:
            # build_data_driven_analysis without API key falls back
            # to data-only analysis — no blocking network call
            try:
                analysis = build_data_driven_analysis(risks_out, company)
            except Exception:
                analysis = {}

        path, _ = QFileDialog.getSaveFileName(
            self, "Save PDF Report", "",
            "PDF Files (*.pdf)",
            selectedFilter="PDF Files (*.pdf)")
        if not path:
            return
        if not path.endswith(".pdf"):
            path += ".pdf"

        self._set_status(
            "⏳  Generating PDF report…", Colors.MEDIUM)

        scope = get_organisation_scope() or {}
        # Merge settings organisation profile as fallback for any missing fields
        # so the report is populated even when no AI analysis scope was saved
        settings_org = self._settings or {}
        merged_scope = {
            "organisation_name":  (scope.get("organisation_name")
                                   or settings_org.get("organisation_name")
                                   or company),
            "industry":           (scope.get("industry")
                                   or settings_org.get("industry", "")),
            "organisation_size":  scope.get("organisation_size", ""),
            "assessment_name":    scope.get("assessment_name",
                                            "Cyber Risk Assessment"),
            "assessment_type":    scope.get("assessment_type",
                                            "Internal Risk Assessment"),
            "assessment_objective": scope.get("assessment_objective", ""),
            "assets_in_scope":    scope.get("assets_in_scope", []),
            "business_units":     scope.get("business_units", []),
        }
        # If org name still blank, use the company field from the export bar
        if not merged_scope["organisation_name"]:
            merged_scope["organisation_name"] = company

        self._pdf_thread = QThread()
        self._pdf_worker = PdfWorker(
            analysis, risks_out, company, path, clf, merged_scope)
        self._pdf_worker.moveToThread(self._pdf_thread)
        self._pdf_thread.started.connect(
            self._pdf_worker.run)
        self._pdf_worker.finished.connect(
            self._on_pdf_done)
        self._pdf_worker.error.connect(
            lambda e: self._set_status(
                f"⚠  PDF error: {e}", Colors.CRITICAL))
        self._pdf_worker.finished.connect(
            self._pdf_thread.quit)
        self._pdf_thread.start()

    def _on_pdf_done(self, path: str) -> None:
        self._set_status(
            f"✅  Report saved: {Path(path).name}",
            Colors.SUCCESS_LT)

    def _export_csv(self) -> None:
        from core.database.db import get_risks as _gr
        self._risks_cache = [dict(r) for r in _gr()]
        risks = self._risks_cache
        if not risks:
            QMessageBox.information(
                self, "Export", "No risks to export.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Save CSV Export", "",
            "CSV Files (*.csv)")
        if not path:
            return
        if not path.endswith(".csv"):
            path += ".csv"

        cols = [
            "id","title","description","category",
            "nist_function","nist_category","iso_domain",
            "cia_component","mitre_tactic","cis_control",
            "likelihood","impact","risk_score",
            "inherent_score","residual_score",
            "owner","status","mitigation","review_date",
            "date_identified","source","confidence","priority",
        ]
        try:
            with open(path, "w", newline="",
                      encoding="utf-8") as f:
                w = csv.DictWriter(
                    f, fieldnames=cols,
                    extrasaction="ignore")
                w.writeheader()
                for r in risks:
                    w.writerow(
                        {c: r.get(c, "") for c in cols})
            audit("EXPORT_CSV",
                  detail=(f"{len(risks)} risks → "
                          f"{Path(path).name}"))
            self._set_status(
                f"✅  CSV exported: {Path(path).name}",
                Colors.SUCCESS_LT)
        except Exception as e:
            self._set_status(
                f"⚠  CSV error: {e}", Colors.CRITICAL)

    def _export_backup(self) -> None:
        try:
            dest = backup_database()
            self._set_status(
                f"✅  Backup created: {dest.name}",
                Colors.SUCCESS_LT)
            # Immediately update backup date/status labels
            self._refresh_backup_status()
        except Exception as e:
            self._set_status(
                f"⚠  Backup error: {e}", Colors.CRITICAL)
